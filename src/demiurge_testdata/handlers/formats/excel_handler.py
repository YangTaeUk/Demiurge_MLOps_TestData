"""Excel FormatHandler — openpyxl 기반"""

from __future__ import annotations

import asyncio
import io
from typing import Any

from demiurge_testdata.core.registry import format_registry
from demiurge_testdata.handlers.base import BaseFormatHandler


@format_registry.register("xlsx")
class ExcelFormatHandler(BaseFormatHandler):
    """Excel (.xlsx) 포맷 핸들러.

    openpyxl을 사용하여 Excel 파일을 생성·파싱한다.
    """

    @property
    def format_name(self) -> str:
        return "xlsx"

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return ".xlsx"

    async def encode(self, records: list[dict]) -> bytes:
        def _encode() -> bytes:
            from openpyxl import Workbook

            if not records:
                return b""

            wb = Workbook()
            ws = wb.active
            ws.title = "data"

            # 헤더
            headers = list(records[0].keys())
            ws.append(headers)

            # 데이터 행
            for rec in records:
                ws.append([rec.get(h) for h in headers])

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        return await asyncio.to_thread(_encode)

    async def decode(self, data: bytes) -> list[dict]:
        def _decode() -> list[dict[str, Any]]:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h) for h in rows[0]]
            return [dict(zip(headers, row)) for row in rows[1:]]

        return await asyncio.to_thread(_decode)
